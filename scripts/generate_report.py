#!/usr/bin/env python3
"""Generate IRD-ready GST report from receipt data."""

import argparse
import json
import sys
from datetime import datetime, date
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
except ImportError:
    import subprocess
    venv_dir = Path.home() / ".openclaw" / "data" / "receipt-to-ird" / ".venv"
    if not venv_dir.exists():
        print("Creating venv and installing openpyxl...", file=sys.stderr)
        subprocess.check_call([sys.executable, "-m", "venv", str(venv_dir)])
    pip = venv_dir / "bin" / "pip"
    subprocess.check_call([str(pip), "install", "openpyxl", "-q"])
    # Re-exec with venv python
    venv_python = venv_dir / "bin" / "python3"
    import os
    os.execv(str(venv_python), [str(venv_python)] + sys.argv)
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers


def get_gst_period(dt: date) -> tuple[int, int, int, int]:
    """Return (start_month, end_month, year, due_month) for 2-monthly GST period."""
    m = dt.month
    start = m if m % 2 == 1 else m - 1
    end = start + 1
    year = dt.year
    due_month = end + 2
    due_year = year
    if due_month > 12:
        due_month -= 12
        due_year += 1
    return start, end, year, due_month


def period_label(start_month: int, end_month: int, year: int) -> str:
    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
              "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    return f"{months[start_month-1]}-{months[end_month-1]} {year}"


def filter_by_period(receipts: list, period: str) -> list:
    """Filter receipts by period. 'current' = current 2-month period, or 'YYYY-MM'."""
    if period == "current":
        today = date.today()
        start_m, end_m, year, _ = get_gst_period(today)
        start_date = date(year, start_m, 1)
        if end_m == 12:
            end_date = date(year + 1, 1, 1)
        else:
            end_date = date(year, end_m + 1, 1)
    elif period == "all":
        return receipts
    else:
        # YYYY-MM format — find the 2-month period containing this month
        parts = period.split("-")
        year, month = int(parts[0]), int(parts[1])
        start_m = month if month % 2 == 1 else month - 1
        end_m = start_m + 1
        start_date = date(year, start_m, 1)
        if end_m == 12:
            end_date = date(year + 1, 1, 1)
        else:
            end_date = date(year, end_m + 1, 1)

    filtered = []
    for r in receipts:
        try:
            rd = date.fromisoformat(r["date"])
            if start_date <= rd < end_date:
                filtered.append(r)
        except (ValueError, KeyError):
            continue
    return filtered


HEADER_FONT = Font(bold=True, size=12, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2B579A", end_color="2B579A", fill_type="solid")
TITLE_FONT = Font(bold=True, size=14)
CURRENCY_FMT = '#,##0.00'
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def style_header_row(ws, row: int, cols: int):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER


def add_summary_sheet(wb: Workbook, receipts: list, period_str: str, business_name: str = "", gst_number: str = ""):
    ws = wb.active
    ws.title = "GST Summary"
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 25

    total_incl = sum(r.get("total", 0) for r in receipts)
    total_gst = sum(r.get("gst", 0) for r in receipts)
    total_excl = total_incl - total_gst

    rows = [
        ("Business Name", business_name or "(not set)"),
        ("IRD / GST Number", gst_number or "(not set)"),
        ("", ""),
        ("GST Period", period_str),
        ("Number of Receipts", len(receipts)),
        ("", ""),
        ("Total Purchases (incl GST)", total_incl),
        ("Total GST on Purchases", total_gst),
        ("Total Purchases (excl GST)", total_excl),
    ]

    ws.cell(row=1, column=1, value="GST Return Summary").font = TITLE_FONT
    ws.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}").font = Font(italic=True, color="666666")

    for i, (label, value) in enumerate(rows, start=4):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True)
        cell = ws.cell(row=i, column=2, value=value)
        if isinstance(value, (int, float)) and label:
            cell.number_format = CURRENCY_FMT
        cell.border = THIN_BORDER


def add_receipts_sheet(wb: Workbook, receipts: list):
    ws = wb.create_sheet("All Receipts")
    headers = ["Date", "Merchant", "Category", "Items", "Amount (excl GST)", "GST", "Total"]
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 40
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 12

    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, 1, len(headers))

    sorted_receipts = sorted(receipts, key=lambda r: r.get("date", ""))
    for i, r in enumerate(sorted_receipts, start=2):
        total = r.get("total", 0)
        gst = r.get("gst", 0)
        excl = round(total - gst, 2)
        items_desc = ", ".join(
            item.get("description", "") for item in r.get("items", [])
        ) or r.get("merchant", "")

        ws.cell(row=i, column=1, value=r.get("date", ""))
        ws.cell(row=i, column=2, value=r.get("merchant", ""))
        ws.cell(row=i, column=3, value=r.get("category", "other"))
        ws.cell(row=i, column=4, value=items_desc)
        ws.cell(row=i, column=5, value=excl).number_format = CURRENCY_FMT
        ws.cell(row=i, column=6, value=gst).number_format = CURRENCY_FMT
        ws.cell(row=i, column=7, value=total).number_format = CURRENCY_FMT

        for col in range(1, 8):
            ws.cell(row=i, column=col).border = THIN_BORDER


def add_category_sheet(wb: Workbook, receipts: list):
    ws = wb.create_sheet("By Category")
    headers = ["Category", "Count", "Total (excl GST)", "GST", "Total (incl GST)"]
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 20

    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, 1, len(headers))

    categories: dict[str, dict] = {}
    for r in receipts:
        cat = r.get("category", "other")
        if cat not in categories:
            categories[cat] = {"count": 0, "total": 0, "gst": 0}
        categories[cat]["count"] += 1
        categories[cat]["total"] += r.get("total", 0)
        categories[cat]["gst"] += r.get("gst", 0)

    for i, (cat, data) in enumerate(sorted(categories.items()), start=2):
        excl = round(data["total"] - data["gst"], 2)
        ws.cell(row=i, column=1, value=cat.title())
        ws.cell(row=i, column=2, value=data["count"])
        ws.cell(row=i, column=3, value=excl).number_format = CURRENCY_FMT
        ws.cell(row=i, column=4, value=round(data["gst"], 2)).number_format = CURRENCY_FMT
        ws.cell(row=i, column=5, value=round(data["total"], 2)).number_format = CURRENCY_FMT
        for col in range(1, 6):
            ws.cell(row=i, column=col).border = THIN_BORDER


def add_ird_sheet(wb: Workbook, receipts: list, period_str: str):
    ws = wb.create_sheet("IRD GST101A")
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 20

    # Box 11: total purchases incl GST
    box_11 = round(sum(r.get("total", 0) for r in receipts), 2)
    # Box 12: GST on purchases = Box 11 × 3/23 (official IRD formula)
    box_12 = round(box_11 * 3 / 23, 2)
    # Box 14: Total GST credit = Box 12 + Box 13 (Box 13 = 0 by default)
    box_14 = box_12

    ws.cell(row=1, column=1, value="IRD GST Return (GST101A) Reference").font = TITLE_FONT
    ws.cell(row=2, column=1, value=f"Period: {period_str}").font = Font(italic=True)
    ws.cell(row=3, column=1, value="Note: Box 5 (sales) must be entered from your accounting records").font = Font(italic=True, color="CC0000")

    headers = ["Box", "Description", "Amount"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=5, column=col, value=h)
    style_header_row(ws, 5, 3)

    ird_rows = [
        ("5", "Total sales and income for the period (incl GST and zero-rated supplies)", "— enter from accounts —"),
        ("6", "Zero-rated supplies included in Box 5", 0.00),
        ("7", "Box 5 minus Box 6", "— calculated —"),
        ("8", "Multiply Box 7 by three and divide by twenty-three (× 3/23)", "— calculated —"),
        ("9", "Adjustments from your calculation sheet", 0.00),
        ("10", "Total GST collected on sales and income (Box 8 + Box 9)", "— calculated —"),
        ("", "", ""),
        ("11", "Total purchases and expenses (incl GST), excl imported goods", box_11),
        ("12", "Multiply Box 11 by three and divide by twenty-three (× 3/23)", box_12),
        ("13", "Credit adjustments from your calculation sheet", 0.00),
        ("14", "Total GST credit for purchases and expenses (Box 12 + Box 13)", box_14),
        ("", "", ""),
        ("15", "Difference (Box 10 minus Box 14). Positive = pay, Negative = refund", "— calculated —"),
    ]

    for i, (box, desc, amt) in enumerate(ird_rows, start=6):
        ws.cell(row=i, column=1, value=box).font = Font(bold=True)
        ws.cell(row=i, column=2, value=desc)
        cell = ws.cell(row=i, column=3, value=amt)
        if isinstance(amt, (int, float)):
            cell.number_format = CURRENCY_FMT
            cell.font = Font(bold=True)
        for col in range(1, 4):
            ws.cell(row=i, column=col).border = THIN_BORDER


def main():
    parser = argparse.ArgumentParser(description="Generate IRD GST report")
    parser.add_argument("--data", required=True, help="Path to receipts.json")
    parser.add_argument("--output", required=True, help="Output XLSX path")
    parser.add_argument("--period", default="current", help="Period: 'current', 'all', or 'YYYY-MM'")
    parser.add_argument("--business-name", default="", help="Business name")
    parser.add_argument("--gst-number", default="", help="GST/IRD number")
    args = parser.parse_args()

    data_path = Path(args.data)
    if not data_path.exists():
        print(f"No receipt data found at {data_path}", file=sys.stderr)
        sys.exit(1)

    with open(data_path) as f:
        all_receipts = json.load(f)

    receipts = filter_by_period(all_receipts, args.period)

    if not receipts:
        print(f"No receipts found for period: {args.period}", file=sys.stderr)
        sys.exit(1)

    # Determine period label
    if args.period == "current":
        today = date.today()
        s, e, y, _ = get_gst_period(today)
        p_label = period_label(s, e, y)
    elif args.period == "all":
        p_label = "All Periods"
    else:
        parts = args.period.split("-")
        year, month = int(parts[0]), int(parts[1])
        s = month if month % 2 == 1 else month - 1
        p_label = period_label(s, s + 1, year)

    wb = Workbook()
    add_summary_sheet(wb, receipts, p_label, args.business_name, args.gst_number)
    add_receipts_sheet(wb, receipts)
    add_category_sheet(wb, receipts)
    add_ird_sheet(wb, receipts, p_label)

    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))

    total_incl = sum(r.get("total", 0) for r in receipts)
    total_gst = sum(r.get("gst", 0) for r in receipts)
    print(json.dumps({
        "period": p_label,
        "receipt_count": len(receipts),
        "total_incl_gst": round(total_incl, 2),
        "total_gst": round(total_gst, 2),
        "total_excl_gst": round(total_incl - total_gst, 2),
        "output": str(output_path),
    }))


if __name__ == "__main__":
    main()
